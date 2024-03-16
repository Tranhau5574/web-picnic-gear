import openpyxl

def read_excel_column(file_path, column, start_row, end_row):
    # Mở tệp Excel
    workbook = openpyxl.load_workbook(file_path)
    
    # Chọn sheet bạn muốn đọc
    sheet = workbook.active
    
    # List để lưu trữ dữ liệu từ cột được chọn
    column_data = []
    
    # Lặp qua từng hàng từ start_row đến end_row và lấy dữ liệu từ cột được chọn
    for row_num in range(start_row, end_row + 1):
        cell_value = sheet[column + str(row_num)].value
        column_data.append(cell_value)
    
    return column_data

def create_excel_file_from_data(data, output_file_path):
    # Tạo một workbook mới
    workbook = openpyxl.Workbook()
    
    # Chọn sheet active
    sheet = workbook.active
    
    # Lặp qua dữ liệu và điền vào ô tương ứng
    row_index = 1
    column_index = 1
    for value in data:
        sheet.cell(row=row_index, column=column_index).value = value
        column_index += 1
        if column_index > 5:  # Nếu đạt đến cột D, điều chỉnh lại cột và tăng dòng
            column_index = 1
            row_index += 1
    
    # Lưu workbook vào file
    workbook.save(output_file_path)

# Thay đổi đường dẫn của tệp Excel của bạn
excel_file_path = "C:\\Users\\tranv\\Downloads\\DS NHÁP.xlsx"

# Cột bạn muốn đọc
column_to_read = 'A'

# Hàng bắt đầu và kết thúc
start_row = 12
end_row = 30

# Đọc dữ liệu từ tệp Excel
column_data = read_excel_column(excel_file_path, column_to_read, start_row, end_row)

# Đường dẫn đến tệp Excel bạn muốn tạo
output_excel_file_path = "C:\\Users\\tranv\\Downloads\\Aha.xlsx"

# Tạo tệp Excel mới từ dữ liệu đã đọc
create_excel_file_from_data(column_data, output_excel_file_path)
